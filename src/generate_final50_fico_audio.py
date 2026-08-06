#!/usr/bin/env python3
"""
Final 50 FICO panel interview audio from Downloads\\Final 50.xlsx

- 3 interviewers: 1 woman + 2 men (rotating, natural panel style)
- Conversational / cornering tone (not dry recitation)
- 25 seconds silence between questions for candidate answers
- Output: C:\\Users\\King2\\Downloads\\final50fico.mp3 (+ companion .txt)

Usage:
  venv\\Scripts\\python.exe generate_final50_fico_audio.py
"""
from __future__ import annotations

import asyncio
import json
import os
import random
import re
import tempfile
from pathlib import Path

from openpyxl import load_workbook

GAP_SECONDS = 25
XLSX = Path(r"C:\Users\King2\Downloads\Final 50.xlsx")
OUT_DIR = Path(r"C:\Users\King2\Downloads")
OUT_MP3 = OUT_DIR / "final50fico.mp3"
OUT_WAV = OUT_DIR / "final50fico.wav"
OUT_TXT = OUT_DIR / "final50fico_questions.txt"

# Panel: one lady, two gentlemen — edge-tts + OpenAI voice maps
INTERVIEWERS = [
    {
        "id": "maya",
        "name": "Maya",
        "role": "Finance Process Lead",
        "gender": "f",
        "edge": "en-US-JennyNeural",
        "openai": "nova",
        "rate": "-4%",
    },
    {
        "id": "daniel",
        "name": "Daniel",
        "role": "SAP Architecture Lead",
        "gender": "m",
        "edge": "en-US-GuyNeural",
        "openai": "onyx",
        "rate": "-8%",
    },
    {
        "id": "marcus",
        "name": "Marcus",
        "role": "Controller / Stakeholder",
        "gender": "m",
        "edge": "en-US-ChristopherNeural",
        "openai": "echo",
        "rate": "-6%",
    },
]


def _load_env() -> None:
    for p in (
        Path(__file__).resolve().parent / ".env",
        Path(r"C:\Users\King2\Desktop\Astra\src\.env"),
    ):
        if not p.exists():
            continue
        for line in p.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))
        break
    if not (os.environ.get("OPENAI_BASE_URL") or "").strip():
        os.environ.pop("OPENAI_BASE_URL", None)


def load_questions(path: Path) -> list[str]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    qs: list[str] = []
    for row in ws.iter_rows(values_only=True):
        q = row[1] if row and len(row) > 1 else None
        if q is None:
            continue
        text = str(q).strip()
        if not text or text.lower().startswith("question"):
            continue
        # collapse whitespace for cleaner speech
        text = re.sub(r"\s+", " ", text)
        qs.append(text)
    return qs


def tts_friendly(q: str) -> str:
    """Light pronunciation fixes for SAP acronyms (speech-only)."""
    repl = [
        (r"\bSAP\b", "S A P"),
        (r"\bFICO\b", "F I C O"),
        (r"\bFI-MM\b", "F I M M"),
        (r"\bFI-SD\b", "F I S D"),
        (r"\bP2P\b", "procure to pay"),
        (r"\bO2C\b", "order to cash"),
        (r"\bOTC\b", "order to cash"),
        (r"\bMTO\b", "make to order"),
        (r"\bAUC\b", "assets under construction"),
        (r"\bWBS\b", "W B S"),
        (r"\bPS\b", "project system"),
        (r"\bCOPA\b", "C O P A"),
        (r"\bCO-PA\b", "C O P A"),
        (r"\bGR/IR\b", "G R I R"),
        (r"\bEBS\b", "electronic bank statement"),
        (r"\bAPP\b", "automatic payment program"),
        (r"\bGL\b", "G L"),
        (r"\bG/L\b", "G L"),
        (r"\bECC\b", "E C C"),
        (r"\bS/4HANA\b", "S four HANA"),
        (r"\bS4hana\b", "S four HANA"),
        (r"\bS4 HANA\b", "S four HANA"),
        (r"\bRAR\b", "revenue accounting and reporting"),
        (r"\bCFIN\b", "central finance"),
        (r"\bMDG\b", "master data governance"),
        (r"\bICMR\b", "intercompany matching and reconciliation"),
        (r"\bFSCM\b", "financial supply chain management"),
        (r"\bTRM\b", "treasury and risk management"),
        (r"\bVIM\b", "vendor invoice management"),
        (r"\bRICEFW\b", "RICEFW"),
        (r"\bASAP\b", "A S A P"),
        (r"\bMRP\b", "M R P"),
        (r"\bML\b", "material ledger"),
        (r"\bIdoc\b", "I-doc"),
        (r"\bIDOC\b", "I-doc"),
        (r"\bBDC\b", "B D C"),
        (r"\bLSMW\b", "L S M W"),
        (r"\bGAAP\b", "GAAP"),
        (r"\bIFRS\b", "I F R S"),
    ]
    out = q
    for pat, rep in repl:
        out = re.sub(pat, rep, out, flags=re.IGNORECASE)
    return out


def conversationalize(q: str, speaker: dict, idx: int, total: int) -> str:
    """
    Turn a dry bank question into a panel-style spoken prompt.
    Cornering: force specificity, evidence, trade-offs — not textbook dumps.
    """
    q_clean = tts_friendly(q.rstrip("?").strip())
    name = speaker["name"]

    # Persona-flavored openers (rotated by index for variety)
    maya_open = [
        f"Maya here. I need a crisp, practical answer — not a theory lecture.",
        f"Maya again. I'm going to hold you to real project evidence.",
        f"Maya. Let's make this uncomfortable in a good way — show me how you think under pressure.",
        f"Maya speaking. I want sequence, ownership, and what you would refuse to sign.",
    ]
    daniel_open = [
        f"Daniel. Architecture lens — I care about design trade-offs and integration risk.",
        f"This is Daniel. Walk me through the system behavior, not the brochure.",
        f"Daniel here. If this breaks at month-end, who owns it and how do you prove it?",
        f"Daniel. Give me the configuration spine and the failure path.",
    ]
    marcus_open = [
        f"Marcus from the controller side. I need audit language and close impact.",
        f"Marcus. Pretend finance leadership is in the room — no fluff.",
        f"Marcus here. If numbers don't tie, what do you block and what do you escalate?",
        f"Marcus. Corner this with controls: detection, prevention, evidence.",
    ]
    opens = {
        "maya": maya_open,
        "daniel": daniel_open,
        "marcus": marcus_open,
    }[speaker["id"]]
    opener = opens[idx % len(opens)]

    # Mid-interview intensity
    if idx == 0:
        bridge = "First question out of the gate."
    elif idx < 3:
        bridge = "Building on how you frame fundamentals."
    elif idx == total // 2:
        bridge = "We're at the midpoint — don't coast."
    elif idx >= total - 3:
        bridge = "Late-stage panel question — stay sharp."
    else:
        bridge = random.choice(
            [
                "Next.",
                "Alright, next topic.",
                "Switching gears.",
                "I'm not done with you yet.",
                "Let's pressure-test this area.",
                "Same panel, harder angle.",
            ]
        )

    # Cornering closers
    closers = [
        "Be specific. Name transactions, objects, and ownership.",
        "If you only give me definitions, I'll push back — give me a real scenario.",
        "Tell me what breaks first if you get this wrong in production.",
        "I want your recommendation and what you would not accept.",
        "Ground this in a project you actually owned.",
        "Clock is on you — structure the answer, then defend it.",
        "Don't hide behind 'it depends' without naming the decision criteria.",
        "If audit walks in tomorrow, what evidence do you show?",
    ]
    closer = closers[idx % len(closers)]

    # Some questions already multi-part — keep as one spoken block
    spoken = (
        f"{opener} {bridge} "
        f"{q_clean}? "
        f"{closer}"
    )
    return re.sub(r"\s+", " ", spoken).strip()


def build_intro() -> list[tuple[dict, str]]:
    """Panel intro lines, each with a speaker."""
    return [
        (
            INTERVIEWERS[1],
            "Good morning. Thanks for joining. This is a panel interview for an S A P F I C O role. "
            "I'm Daniel, architecture lead. I'll focus on design, integration, and what fails under load.",
        ),
        (
            INTERVIEWERS[0],
            "I'm Maya, finance process lead. I'll challenge how you run procure to pay, order to cash, close, "
            "and whether your answers sound like real delivery or slides.",
        ),
        (
            INTERVIEWERS[2],
            "And I'm Marcus, representing controllers and stakeholders. I care about controls, audit trail, "
            "and whether the books can close when something is off. We'll rotate questions. "
            "After each question you'll have about twenty-five seconds of silence to answer out loud before we continue. "
            "Let's begin.",
        ),
    ]


def build_outro() -> list[tuple[dict, str]]:
    return [
        (
            INTERVIEWERS[0],
            "Maya here — that's the end of the question set. Strong candidates gave sequences, ownership, and trade-offs.",
        ),
        (
            INTERVIEWERS[1],
            "Daniel — if you want a second pass, re-run this file and answer tighter where you hedged.",
        ),
        (
            INTERVIEWERS[2],
            "Marcus — thank you. Session complete.",
        ),
    ]


def main() -> int:
    from pydub import AudioSegment
    from pydub.generators import Sine

    _load_env()
    random.seed(50)  # stable variety

    if not XLSX.exists():
        print(f"Missing Excel: {XLSX}", flush=True)
        return 1

    questions = load_questions(XLSX)
    if not questions:
        print("No questions found in Excel.", flush=True)
        return 1

    api_key = (os.environ.get("OPENAI_API_KEY") or "").strip()
    use_openai = bool(api_key) and len(api_key) > 20 and api_key.startswith("sk-")
    engine = "openai" if use_openai else "edge-tts"
    print(f"TTS engine: {engine}", flush=True)
    print(f"Questions from Excel: {len(questions)}", flush=True)
    print(f"Gap between questions: {GAP_SECONDS}s", flush=True)
    print(f"Output: {OUT_MP3}", flush=True)

    # Assign speakers in panel rotation with slight shuffle blocks (still balanced)
    speaker_cycle = [INTERVIEWERS[0], INTERVIEWERS[1], INTERVIEWERS[2]]
    # Prefer: Maya, Daniel, Marcus, Daniel, Maya, Marcus... not pure round-robin monotony
    pattern = [0, 1, 2, 1, 0, 2, 1, 0, 2, 1]
    assignments: list[dict] = []
    for i in range(len(questions)):
        assignments.append(speaker_cycle[pattern[i % len(pattern)]])

    # Transcript
    lines = [
        "Final 50 FICO panel interview (from Final 50.xlsx)",
        f"Questions: {len(questions)}",
        f"Gap: {GAP_SECONDS}s between questions",
        f"Interviewers: Maya (F), Daniel (M), Marcus (M)",
        f"TTS: {engine}",
        "",
        "=== INTRO ===",
    ]
    for sp, text in build_intro():
        lines.append(f"[{sp['name']} / {sp['role']}] {text}")
    lines.append("")
    lines.append("=== QUESTIONS ===")
    spoken_qs: list[tuple[dict, str]] = []
    for i, q in enumerate(questions):
        sp = assignments[i]
        spoken = conversationalize(q, sp, i, len(questions))
        spoken_qs.append((sp, spoken))
        lines.append(f"Q{i+1:02d} [{sp['name']}]")
        lines.append(f"  SOURCE: {q}")
        lines.append(f"  SPOKEN: {spoken}")
        lines.append("")
    lines.append("=== OUTRO ===")
    for sp, text in build_outro():
        lines.append(f"[{sp['name']}] {text}")
    OUT_TXT.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {OUT_TXT}", flush=True)

    openai_client = None
    if use_openai:
        from openai import OpenAI

        openai_client = OpenAI(api_key=api_key)

    soft_beep = Sine(740).to_audio_segment(duration=60).apply_gain(-16)
    hard_beep = (
        Sine(620).to_audio_segment(duration=55).apply_gain(-13)
        + AudioSegment.silent(duration=40)
        + Sine(920).to_audio_segment(duration=55).apply_gain(-13)
    )
    gap = AudioSegment.silent(duration=GAP_SECONDS * 1000)
    short = AudioSegment.silent(duration=450)
    tmp_dir = Path(tempfile.mkdtemp(prefix="final50fico_tts_"))

    def tts(text: str, speaker: dict, label: str) -> AudioSegment:
        print(f"  TTS [{speaker['name']}]: {label}...", flush=True)
        tmp = tmp_dir / f"{label}.mp3"
        if use_openai and openai_client is not None:
            with openai_client.audio.speech.with_streaming_response.create(
                model="tts-1-hd",
                voice=speaker["openai"],
                input=text,
            ) as resp:
                resp.stream_to_file(str(tmp))
        else:
            import edge_tts

            communicate = edge_tts.Communicate(
                text,
                voice=speaker["edge"],
                rate=speaker.get("rate", "-5%"),
            )
            asyncio.run(communicate.save(str(tmp)))
        return AudioSegment.from_file(tmp)

    print("Generating panel audio...", flush=True)
    combined = AudioSegment.silent(duration=600)

    for i, (sp, text) in enumerate(build_intro()):
        combined += tts(text, sp, f"intro_{i}_{sp['id']}") + short

    combined += AudioSegment.silent(duration=700)

    n = len(spoken_qs)
    for i, (sp, text) in enumerate(spoken_qs, 1):
        cue = hard_beep if i == 1 or i % 5 == 0 else soft_beep
        combined += cue + AudioSegment.silent(duration=180)
        combined += tts(text, sp, f"Q{i:02d}_{sp['id']}")
        if i < n:
            print(f"  gap {GAP_SECONDS}s after Q{i}/{n}", flush=True)
            combined += gap
        else:
            combined += AudioSegment.silent(duration=1200)

    combined += short
    for i, (sp, text) in enumerate(build_outro()):
        combined += tts(text, sp, f"outro_{i}_{sp['id']}") + short
    combined += AudioSegment.silent(duration=800)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"Exporting {OUT_MP3} ...", flush=True)
    combined.export(str(OUT_MP3), format="mp3", bitrate="192k")
    duration_s = len(combined) / 1000.0
    print(f"Wrote {OUT_MP3} ({duration_s:.1f}s / {duration_s/60:.1f} min)", flush=True)

    try:
        print(f"Also exporting {OUT_WAV} ...", flush=True)
        combined.export(str(OUT_WAV), format="wav")
        print(f"Wrote {OUT_WAV}", flush=True)
    except Exception as e:
        print(f"WAV export skipped: {e}", flush=True)

    meta = (
        f"\n---\n"
        f"Questions spoken: {n}\n"
        f"Answer window: {GAP_SECONDS}s between questions\n"
        f"Total duration: {duration_s/60:.1f} minutes\n"
        f"Interviewers: Maya (female), Daniel (male), Marcus (male)\n"
        f"Source: {XLSX}\n"
        f"Engine: {engine}\n"
    )
    with OUT_TXT.open("a", encoding="utf-8") as f:
        f.write(meta)
    print(meta, flush=True)
    print("Done.", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
