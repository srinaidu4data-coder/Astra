"""Export Final 50.xlsx questions + spoken lines into frontend ready-made mock JSON."""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

XLSX = Path(r"C:\Users\King2\Downloads\Final 50.xlsx")
TXT = Path(r"C:\Users\King2\Downloads\final50fico_questions.txt")
OUT = Path(
    r"C:\Users\King2\Desktop\Astra\interview-pulse-ai\src\data\readyMadeMocks.generated.json"
)

wb = load_workbook(XLSX, data_only=True)
ws = wb.active
sources: list[str] = []
for row in ws.iter_rows(values_only=True):
    q = row[1] if row and len(row) > 1 else None
    if q is None:
        continue
    text = re.sub(r"\s+", " ", str(q).strip())
    if not text or text.lower().startswith("question"):
        continue
    sources.append(text)

spoken_by_n: dict[int, dict] = {}
if TXT.exists():
    t = TXT.read_text(encoding="utf-8")
    for m in re.finditer(
        r"Q(\d+)\s+\[([^\]]+)\]\s*\n\s*SOURCE:\s*(.+)\n\s*SPOKEN:\s*(.+)",
        t,
    ):
        n = int(m.group(1))
        spoken_by_n[n] = {
            "speaker": m.group(2).strip(),
            "source": m.group(3).strip(),
            "spoken": m.group(4).strip(),
        }

questions = []
for i, src in enumerate(sources, 1):
    meta = spoken_by_n.get(i, {})
    questions.append(
        {
            "id": f"fico_final50_q{i:02d}",
            "text": src,
            "spoken_text": meta.get("spoken") or src,
            "category": "sap-fico",
            "hint": f"Panel interviewer: {meta.get('speaker', 'panel')}",
            "bridge": "",
            "speaker": meta.get("speaker", "Panel"),
        }
    )

intro = (
    "Good morning. Thanks for joining. This is a panel-style mock for SAP FICO. "
    "I'm Daniel, architecture lead. Maya will push process depth, and Marcus will "
    "pressure controls and close impact. After each question, answer out loud — "
    "we score STAR, depth, and communication. Let's begin."
)
closing = (
    "That wraps the SAP FICO Final 50 panel set. Review your debrief for strengths, "
    "gaps, and a practice plan. You can re-run any ready-made mock from the Mock tab."
)

payload = {
    "categories": [
        {
            "id": "sap-fico",
            "label": "SAP FICO",
            "description": "Finance & Controlling panel mocks — enterprise structure through ICMR, FSCM, and close.",
            "packs": [
                {
                    "id": "sap-fico-final-50",
                    "title": "Final 50 — FICO Panel",
                    "subtitle": "41 hard panel questions · Maya · Daniel · Marcus style",
                    "difficulty": "hard",
                    "focus": "technical",
                    "persona": "strict-tech-lead",
                    "job_title": "SAP FICO Consultant",
                    "company": "",
                    "question_count": len(questions),
                    "answer_seconds": 90,
                    "tags": ["SAP", "FICO", "S/4HANA", "Panel", "Hard"],
                    "audio_url": "/mocks/sap-fico/final50fico.mp3",
                    "intro_script": intro,
                    "closing_script": closing,
                    "questions": questions,
                }
            ],
        }
    ]
}

OUT.parent.mkdir(parents=True, exist_ok=True)
OUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
print(f"Wrote {OUT} with {len(questions)} questions")
